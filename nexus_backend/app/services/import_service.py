"""
数据导入服务 - CSV/Excel 解析和批量导入
支持员工和客户数据的批量导入，包含数据校验和错误处理
"""
import csv
import io
import re
from typing import List, Dict, Any, Optional
import logging

# Try to import openpyxl, but don't fail if not available
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ImportService:
    """数据导入服务类，处理 CSV 和 Excel 文件的解析和导入"""
    
    @staticmethod
    def _parse_file(contents: bytes, filename: str) -> List[Dict[str, Any]]:
        """
        解析上传的文件内容（CSV 或 Excel）
        
        Args:
            contents: 文件二进制内容
            filename: 文件名（用于判断文件类型）
        
        Returns:
            解析后的数据列表，每行为一个字典
        
        Raises:
            ValueError: 文件格式不支持或解析失败
        """
        file_ext = filename.lower().split('.')[-1]
        
        if file_ext == 'csv':
            return ImportService._parse_csv(contents)
        elif file_ext in ['xlsx', 'xls']:
            return ImportService._parse_excel(contents, file_ext)
        else:
            raise ValueError(f"不支持的文件格式: {file_ext}")
    
    @staticmethod
    def _parse_csv(contents: bytes) -> List[Dict[str, Any]]:
        """
        解析 CSV 文件，自动检测编码（utf-8 或 gbk）
        
        Args:
            contents: CSV 文件二进制内容
        
        Returns:
            解析后的数据列表
        """
        # 尝试 UTF-8 编码
        try:
            text = contents.decode('utf-8')
        except UnicodeDecodeError:
            # 如果 UTF-8 失败，尝试 GBK
            try:
                text = contents.decode('gbk')
            except UnicodeDecodeError:
                raise ValueError("无法识别文件编码，请使用 UTF-8 或 GBK 编码")
        
        # 使用 csv.DictReader 解析
        text_io = io.StringIO(text)
        reader = csv.DictReader(text_io)
        
        data = []
        for row in reader:
            # 过滤空行
            if any(row.values()):
                data.append(row)
        
        return data
    
    @staticmethod
    def _parse_excel(contents: bytes, file_ext: str) -> List[Dict[str, Any]]:
        """
        解析 Excel 文件
        
        Args:
            contents: Excel 文件二进制内容
            file_ext: 文件扩展名
        
        Returns:
            解析后的数据列表
        """
        if not OPENPYXL_AVAILABLE:
            raise ValueError("Excel 文件解析需要 openpyxl 库，请联系管理员安装")
        
        # 使用 BytesIO 创建文件对象
        file_io = io.BytesIO(contents)
        
        try:
            # 使用只读模式加载工作簿
            workbook = load_workbook(file_io, read_only=True)
            sheet = workbook.active
            
            # 获取表头（第一行）
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            
            headers = rows[0]
            data = []
            
            # 解析数据行
            for row in rows[1:]:
                # 过滤空行
                if any(row):
                    row_dict = {str(headers[i]): row[i] for i in range(len(headers)) if i < len(row)}
                    data.append(row_dict)
            
            return data
        except Exception as e:
            logger.error(f"Excel 解析失败: {e}")
            raise ValueError(f"Excel 文件解析失败: {str(e)}")
    
    @staticmethod
    async def preview(contents: bytes, filename: str) -> Dict[str, Any]:
        """
        预览导入数据（返回前10行和总行数）
        
        Args:
            contents: 文件二进制内容
            filename: 文件名
        
        Returns:
            包含 rows（前10行）和 total（总行数）的字典
        """
        try:
            data = ImportService._parse_file(contents, filename)
            return {
                "rows": data[:10],
                "total": len(data)
            }
        except Exception as e:
            logger.error(f"文件预览失败: {e}")
            raise ValueError(f"文件预览失败: {str(e)}")
    
    @staticmethod
    def get_template(template_type: str) -> str:
        """
        获取导入模板（CSV 格式）
        
        Args:
            template_type: 模板类型（employees 或 customers）
        
        Returns:
            CSV 格式的模板字符串
        """
        templates = {
            "employees": "姓名,邮箱,部门,角色,手机号",
            "customers": "客户名称,联系人,联系电话,邮箱,公司,来源,备注"
        }
        
        template = templates.get(template_type)
        if not template:
            raise ValueError(f"不支持的模板类型: {template_type}")
        
        return template
    
    @staticmethod
    def _validate_email(email: str) -> bool:
        """验证邮箱格式"""
        if not email:
            return False
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))
    
    @staticmethod
    def _validate_phone(phone: str) -> bool:
        """验证中国手机号格式（1开头11位）"""
        if not phone:
            return False
        pattern = r'^1\d{10}$'
        return bool(re.match(pattern, phone))
    
    @staticmethod
    def _validate_employee(row: Dict[str, Any], row_num: int) -> Optional[Dict[str, Any]]:
        """
        验证员工数据
        
        Args:
            row: 员工数据字典
            row_num: 行号（用于错误报告）
        
        Returns:
            如果验证失败，返回错误字典；否则返回 None
        """
        # 角色值校验
        role = row.get("role", "").strip()
        if role and role not in ["employee", "manager", "boss"]:
            return {
                "row": row_num,
                "data": row,
                "reason": f"角色值无效，只允许 employee/manager/boss，当前值: {role}",
                "field": "role"
            }
        
        # 手机号格式校验
        phone = row.get("phone", "").strip()
        if phone and not ImportService._validate_phone(phone):
            return {
                "row": row_num,
                "data": row,
                "reason": f"手机号格式无效，应为1开头的11位数字，当前值: {phone}",
                "field": "phone"
            }
        
        # 部门非空校验
        department = row.get("department", "").strip()
        if not department:
            return {
                "row": row_num,
                "data": row,
                "reason": "部门不能为空",
                "field": "department"
            }
        
        return None
    
    @staticmethod
    def _validate_customer(row: Dict[str, Any], row_num: int) -> Optional[Dict[str, Any]]:
        """
        验证客户数据
        
        Args:
            row: 客户数据字典
            row_num: 行号（用于错误报告）
        
        Returns:
            如果验证失败，返回错误字典；否则返回 None
        """
        # 联系电话格式校验
        phone = row.get("phone", "").strip()
        if phone and not ImportService._validate_phone(phone):
            return {
                "row": row_num,
                "data": row,
                "reason": f"联系电话格式无效，应为1开头的11位数字，当前值: {phone}",
                "field": "phone"
            }
        
        # 来源字段校验和标准化
        source = row.get("source", "").strip()
        allowed_sources = ["推荐", "网络", "电话", "展会", "其他"]
        if source and source not in allowed_sources:
            # 不在范围的标为"其他"
            row["source"] = "其他"
        
        return None
    
    @staticmethod
    def _detect_duplicates(rows: List[Dict[str, Any]], key_fields: List[str]) -> List[Dict[str, Any]]:
        """
        检测文件内部重复行
        
        Args:
            rows: 数据行列表
            key_fields: 用于检测重复的关键字段列表
        
        Returns:
            重复行的错误信息列表
        """
        seen = {}
        duplicates = []
        
        for idx, row in enumerate(rows, start=2):  # 从第2行开始（第1行是表头）
            # 构建复合键
            key_values = []
            for field in key_fields:
                value = row.get(field, "").strip()
                if value:  # 只考虑非空值
                    key_values.append(value)
            
            if not key_values:
                continue
            
            key = tuple(key_values)
            
            if key in seen:
                duplicates.append({
                    "row": idx,
                    "data": row,
                    "reason": f"文件内部重复：与第{seen[key]}行的 {'/'.join(key_fields)} 字段值相同 ({', '.join(key_values)})",
                    "field": "/".join(key_fields)
                })
            else:
                seen[key] = idx
        
        return duplicates
    
    @staticmethod
    async def import_employees(
        contents: bytes,
        filename: str,
        user_id: str,
        db_client: Any,
        mode: str = "insert"
    ) -> Dict[str, Any]:
        """
        导入员工数据
        
        Args:
            contents: 文件二进制内容
            filename: 文件名
            user_id: 当前操作用户ID
            db_client: 数据库客户端（支持 RLS）
            mode: 导入模式（insert 或 incremental）
        
        Returns:
            导入结果统计: success_count, skip_count, error_count, errors
        """
        try:
            data = ImportService._parse_file(contents, filename)
        except Exception as e:
            logger.error(f"文件解析失败: {e}")
            return {
                "success_count": 0,
                "skip_count": 0,
                "error_count": 0,
                "errors": [{"row": 0, "data": {}, "reason": f"文件解析失败: {str(e)}", "field": "file"}]
            }
        
        success_count = 0
        skip_count = 0
        error_count = 0
        errors = []
        
        # 字段映射（兼容中英文表头）
        field_map = {
            "姓名": "name",
            "name": "name",
            "邮箱": "email",
            "email": "email",
            "部门": "department",
            "department": "department",
            "角色": "role",
            "role": "role",
            "手机号": "phone",
            "phone": "phone"
        }
        
        # 标准化所有数据行
        normalized_data = []
        for row in data:
            normalized_row = {}
            for k, v in row.items():
                if k in field_map:
                    normalized_row[field_map[k]] = v
            normalized_data.append(normalized_row)
        
        # C5: 检测文件内部重复（基于邮箱）
        duplicates = ImportService._detect_duplicates(normalized_data, ["email"])
        errors.extend(duplicates)
        error_count += len(duplicates)
        
        for idx, row in enumerate(normalized_data, start=2):  # 从第2行开始（第1行是表头）
            try:
                # 校验必填字段
                name = row.get("name", "").strip()
                email = row.get("email", "").strip()
                
                if not name:
                    errors.append({
                        "row": idx,
                        "data": row,
                        "reason": "姓名不能为空",
                        "field": "name"
                    })
                    error_count += 1
                    continue
                
                if not email:
                    errors.append({
                        "row": idx,
                        "data": row,
                        "reason": "邮箱不能为空",
                        "field": "email"
                    })
                    error_count += 1
                    continue
                
                # 校验邮箱格式
                if not ImportService._validate_email(email):
                    errors.append({
                        "row": idx,
                        "data": row,
                        "reason": f"邮箱格式无效: {email}",
                        "field": "email"
                    })
                    error_count += 1
                    continue
                
                # C3: 调用 _validate_employee 进行额外校验
                validation_error = ImportService._validate_employee(row, idx)
                if validation_error:
                    errors.append(validation_error)
                    error_count += 1
                    continue
                
                # 检查邮箱是否已存在
                existing = await db_client.table("users")\
                    .select("id")\
                    .eq("email", email)\
                    .maybe_single()\
                    .execute()
                
                # 构建用户数据
                user_data = {
                    "name": name,
                    "email": email,
                    "department": row.get("department", "").strip() or None,
                    "role": row.get("role", "employee").strip(),  # 默认为 employee
                    "phone": row.get("phone", "").strip() or None,
                    "password": "Nexus@123",  # 默认密码
                    "score": 0,
                    "total_bonus": 0
                }
                
                # 根据 mode 选择 insert 或 upsert
                if mode == "incremental":
                    # 增量模式：使用 upsert
                    if existing.data:
                        # 更新已有记录
                        await db_client.table("users")\
                            .update(user_data)\
                            .eq("id", existing.data["id"])\
                            .execute()
                        success_count += 1
                    else:
                        # 新增记录
                        await db_client.table("users").insert(user_data).execute()
                        success_count += 1
                else:
                    # 默认 insert 模式：跳过已存在的记录
                    if existing.data:
                        skip_count += 1
                        continue
                    
                    # 插入数据库
                    await db_client.table("users").insert(user_data).execute()
                    success_count += 1
                
            except Exception as e:
                logger.error(f"第{idx}行导入失败: {e}")
                errors.append({
                    "row": idx,
                    "data": row,
                    "reason": str(e),
                    "field": "unknown"
                })
                error_count += 1
        
        return {
            "success_count": success_count,
            "skip_count": skip_count,
            "error_count": error_count,
            "errors": errors[:20]  # 最多返回前20条错误
        }
    
    @staticmethod
    async def import_customers(
        contents: bytes,
        filename: str,
        user_id: str,
        db_client: Any,
        mode: str = "insert"
    ) -> Dict[str, Any]:
        """
        导入客户数据
        
        Args:
            contents: 文件二进制内容
            filename: 文件名
            user_id: 当前操作用户ID
            db_client: 数据库客户端（支持 RLS）
            mode: 导入模式（insert 或 incremental）
        
        Returns:
            导入结果统计: success_count, skip_count, error_count, errors
        """
        try:
            data = ImportService._parse_file(contents, filename)
        except Exception as e:
            logger.error(f"文件解析失败: {e}")
            return {
                "success_count": 0,
                "skip_count": 0,
                "error_count": 0,
                "errors": [{"row": 0, "data": {}, "reason": f"文件解析失败: {str(e)}", "field": "file"}]
            }
        
        success_count = 0
        skip_count = 0
        error_count = 0
        errors = []
        
        # 字段映射（兼容中英文表头）
        field_map = {
            "客户名称": "name",
            "name": "name",
            "联系人": "contact_person",
            "contact_person": "contact_person",
            "联系电话": "phone",
            "phone": "phone",
            "邮箱": "email",
            "email": "email",
            "公司": "company",
            "company": "company",
            "来源": "source",
            "source": "source",
            "备注": "notes",
            "notes": "notes"
        }
        
        # 标准化所有数据行
        normalized_data = []
        for row in data:
            normalized_row = {}
            for k, v in row.items():
                if k in field_map:
                    normalized_row[field_map[k]] = v
            normalized_data.append(normalized_row)
        
        # C5: 检测文件内部重复（基于客户名称和公司）
        duplicates = ImportService._detect_duplicates(normalized_data, ["name", "company"])
        errors.extend(duplicates)
        error_count += len(duplicates)
        
        for idx, row in enumerate(normalized_data, start=2):  # 从第2行开始
            try:
                # 校验必填字段
                name = row.get("name", "").strip()
                
                if not name:
                    errors.append({
                        "row": idx,
                        "data": row,
                        "reason": "客户名称不能为空",
                        "field": "name"
                    })
                    error_count += 1
                    continue
                
                # 校验邮箱格式（如果提供）
                email = row.get("email", "").strip()
                if email and not ImportService._validate_email(email):
                    errors.append({
                        "row": idx,
                        "data": row,
                        "reason": f"邮箱格式无效: {email}",
                        "field": "email"
                    })
                    error_count += 1
                    continue
                
                # C4: 调用 _validate_customer 进行额外校验
                validation_error = ImportService._validate_customer(row, idx)
                if validation_error:
                    errors.append(validation_error)
                    error_count += 1
                    continue
                
                # 检查客户是否已存在（按名称 + 公司）
                company = row.get("company", "").strip()
                query = db_client.table("customers").select("id").eq("name", name)
                
                if company:
                    query = query.eq("company", company)
                
                existing = await query.maybe_single().execute()
                
                # 构建客户数据
                customer_data = {
                    "name": name,
                    "contact_person": row.get("contact_person", "").strip() or None,
                    "phone": row.get("phone", "").strip() or None,
                    "email": email or None,
                    "company": company or None,
                    "source": row.get("source", "").strip() or "批量导入",
                    "notes": row.get("notes", "").strip() or None,
                    "created_by": user_id
                }
                
                # 根据 mode 选择 insert 或 upsert
                if mode == "incremental":
                    # 增量模式：使用 upsert
                    if existing.data:
                        # 更新已有记录
                        await db_client.table("customers")\
                            .update(customer_data)\
                            .eq("id", existing.data["id"])\
                            .execute()
                        success_count += 1
                    else:
                        # 新增记录
                        await db_client.table("customers").insert(customer_data).execute()
                        success_count += 1
                else:
                    # 默认 insert 模式：跳过已存在的记录
                    if existing.data:
                        skip_count += 1
                        continue
                    
                    # 插入数据库
                    await db_client.table("customers").insert(customer_data).execute()
                    success_count += 1
                
            except Exception as e:
                logger.error(f"第{idx}行导入失败: {e}")
                errors.append({
                    "row": idx,
                    "data": row,
                    "reason": str(e),
                    "field": "unknown"
                })
                error_count += 1
        
        return {
            "success_count": success_count,
            "skip_count": skip_count,
            "error_count": error_count,
            "errors": errors[:20]  # 最多返回前20条错误
        }
